import numpy as np
from scipy import stats
import plotly.graph_objects as go


def fit_standard_curve(concentration, absorbance, model_type="linear"):
    try:
        x = np.array(concentration, dtype=float)
        y = np.array(absorbance, dtype=float)

        if model_type == "linear":
            slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
            r_squared = r_value ** 2
            if intercept >= 0:
                eq = "y = {:.4f}x + {:.4f}".format(slope, intercept)
            else:
                eq = "y = {:.4f}x - {:.4f}".format(slope, abs(intercept))
            return {
                "success": True,
                "model": "linear",
                "params": {"slope": slope, "intercept": intercept},
                "r_squared": r_squared,
                "equation": eq,
                "std_err": std_err,
                "p_value": p_value
            }

        elif model_type == "polynomial":
            coeffs = np.polyfit(x, y, 2)
            a, b, c = coeffs
            y_pred = np.polyval(coeffs, x)
            ss_res = np.sum((y - y_pred) ** 2)
            ss_tot = np.sum((y - np.mean(y)) ** 2)
            r_sq = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
            sb = "+" if b >= 0 else "-"
            sc = "+" if c >= 0 else "-"
            eq = "y = {:.4f}x^2 {} {:.4f}x {} {:.4f}".format(
                a, sb, abs(b), sc, abs(c))
            return {
                "success": True,
                "model": "polynomial",
                "params": {"a": a, "b": b, "c": c},
                "r_squared": r_sq,
                "equation": eq
            }

    except Exception as e:
        return {"success": False, "error": str(e)}


def calculate_unknown(absorbance_value, fit_result):
    try:
        y = float(absorbance_value)
        if fit_result["model"] == "linear":
            m = fit_result["params"]["slope"]
            b = fit_result["params"]["intercept"]
            if m == 0:
                return None
            return round((y - b) / m, 4)
        elif fit_result["model"] == "polynomial":
            a = fit_result["params"]["a"]
            b = fit_result["params"]["b"]
            c = fit_result["params"]["c"]
            disc = b**2 - 4*a*(c - y)
            if disc < 0:
                return None
            x1 = (-b + np.sqrt(disc)) / (2*a)
            x2 = (-b - np.sqrt(disc)) / (2*a)
            candidates = [v for v in [x1, x2] if v >= 0]
            return round(min(candidates), 4) if candidates else None
    except Exception:
        return None


def build_standard_curve_figure(concentration, absorbance,
                                 fit_result, assay_name,
                                 x_unit, y_unit, unknowns=None):
    x = np.array(concentration, dtype=float)
    y = np.array(absorbance, dtype=float)
    x_sm = np.linspace(min(x), max(x) * 1.05, 300)

    if fit_result["model"] == "linear":
        m = fit_result["params"]["slope"]
        b = fit_result["params"]["intercept"]
        y_sm = m * x_sm + b
    else:
        a = fit_result["params"]["a"]
        b = fit_result["params"]["b"]
        c = fit_result["params"]["c"]
        y_sm = a * x_sm**2 + b * x_sm + c

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=x_sm, y=y_sm, mode="lines",
        name="Fitted ({})".format(fit_result["model"]),
        line=dict(color="#378ADD", width=2.5)
    ))

    fig.add_trace(go.Scatter(
        x=x, y=y, mode="markers",
        name="Standard points",
        marker=dict(size=10, color="#1D9E75",
                    line=dict(width=1.5, color="white"))
    ))

    if unknowns:
        ux = [u["concentration"] for u in unknowns
              if u["concentration"] is not None]
        uy = [u["absorbance"] for u in unknowns
              if u["concentration"] is not None]
        ul = [u.get("label", "S{}".format(i+1))
              for i, u in enumerate(unknowns)
              if u["concentration"] is not None]
        if ux:
            fig.add_trace(go.Scatter(
                x=ux, y=uy, mode="markers+text",
                name="Unknown samples",
                marker=dict(size=13, color="#f59e0b", symbol="diamond",
                            line=dict(width=1.5, color="white")),
                text=ul, textposition="top center"
            ))

    fig.update_layout(
        title=dict(
            text="Standard Curve — {}  |  {}  |  R2 = {:.4f}".format(
                assay_name,
                fit_result["equation"],
                fit_result["r_squared"]
            ),
            font_size=14
        ),
        xaxis_title="Concentration ({})".format(x_unit),
        yaxis_title="Absorbance ({})".format(y_unit),
        plot_bgcolor="white",
        height=430,
        legend=dict(orientation="h", yanchor="bottom",
                    y=1.02, xanchor="right", x=1),
        margin=dict(l=60, r=20, t=80, b=50)
    )
    fig.update_xaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=True)
    fig.update_yaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=True)
    return fig


def build_excel_standard_chart(concentration, absorbance, fit_result,
                                assay_name, x_unit, y_unit, unknowns=None):
    import io
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, Reference, Series

    x = np.array(concentration, dtype=float)
    y = np.array(absorbance, dtype=float)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Standard Data"
    ws1["A1"] = f"Concentration ({x_unit})"
    ws1["B1"] = f"Absorbance ({y_unit})"
    for i, (c, a) in enumerate(zip(x, y), start=2):
        ws1.cell(i, 1, float(c))
        ws1.cell(i, 2, float(a))

    ws2 = wb.create_sheet("Fitted Curve")
    x_sm = np.linspace(min(x), max(x) * 1.05, 100)
    if fit_result["model"] == "linear":
        m, b = fit_result["params"]["slope"], fit_result["params"]["intercept"]
        y_sm = m * x_sm + b
    else:
        a_, b_, c_ = (fit_result["params"]["a"], fit_result["params"]["b"],
                      fit_result["params"]["c"])
        y_sm = a_ * x_sm ** 2 + b_ * x_sm + c_
    ws2["A1"] = f"Concentration ({x_unit})"
    ws2["B1"] = f"Absorbance ({y_unit}) fitted"
    for i, (c, a) in enumerate(zip(x_sm, y_sm), start=2):
        ws2.cell(i, 1, float(c))
        ws2.cell(i, 2, float(a))

    ws3 = wb.create_sheet("Parameters")
    ws3.append(["Parameter", "Value"])
    ws3.append(["Assay", assay_name])
    ws3.append(["Model", fit_result["model"]])
    ws3.append(["Equation", fit_result["equation"]])
    ws3.append(["R²", round(fit_result["r_squared"], 6)])

    has_unknowns = False
    if unknowns:
        valid = [u for u in unknowns if u["concentration"] is not None]
        if valid:
            has_unknowns = True
            ws4 = wb.create_sheet("Unknown Samples")
            ws4.append(["Sample", f"Absorbance ({y_unit})", f"Concentration ({x_unit})"])
            for u in valid:
                ws4.append([u["label"], u["absorbance"], u["concentration"]])

    n_std = len(x)
    chart = ScatterChart()
    chart.title = f"{assay_name} | {fit_result['equation']} | R²={fit_result['r_squared']:.4f}"
    chart.style = 10
    chart.xAxis.title = f"Concentration ({x_unit})"
    chart.yAxis.title = f"Absorbance ({y_unit})"
    chart.width, chart.height = 24, 14

    x_std = Reference(ws1, min_col=1, min_row=2, max_row=n_std + 1)
    y_std = Reference(ws1, min_col=2, min_row=1, max_row=n_std + 1)
    s_std = Series(y_std, x_std, title_from_data=True)
    s_std.marker.symbol = "circle"
    s_std.marker.size = 8
    s_std.graphicalProperties.line.noFill = True
    chart.series.append(s_std)

    x_fit = Reference(ws2, min_col=1, min_row=2, max_row=101)
    y_fit = Reference(ws2, min_col=2, min_row=1, max_row=101)
    s_fit = Series(y_fit, x_fit, title_from_data=True)
    s_fit.marker.symbol = "none"
    s_fit.graphicalProperties.line.solidFill = "378ADD"
    chart.series.append(s_fit)

    if has_unknowns:
        x_unk = Reference(ws4, min_col=3, min_row=2, max_row=len(valid) + 1)
        y_unk = Reference(ws4, min_col=2, min_row=1, max_row=len(valid) + 1)
        s_unk = Series(y_unk, x_unk, title_from_data=True)
        s_unk.marker.symbol = "diamond"
        s_unk.marker.size = 10
        s_unk.graphicalProperties.line.noFill = True
        chart.series.append(s_unk)

    ws1.add_chart(chart, "D2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ASSAY_PRESETS = {
    "Bradford Protein Assay": {
        "x_unit": "ug/mL",
        "y_unit": "OD595",
        "wavelength": "595 nm",
        "model": "linear",
        "description": (
            "BSA standard curve at 595 nm. Coomassie Brilliant Blue G-250 "
            "binds to protein producing a blue colour proportional to "
            "protein concentration."
        ),
        "sample_data": {
            "concentration": [0, 25, 50, 100, 200, 300, 400, 500],
            "absorbance":    [0.000, 0.112, 0.224, 0.445, 0.871,
                              1.280, 1.651, 1.990]
        }
    },
    "IAA Colorimetric Assay (Salkowski)": {
        "x_unit": "ug/mL",
        "y_unit": "OD530",
        "wavelength": "530 nm",
        "model": "linear",
        "description": (
            "IAA standard curve using Salkowski reagent (FeCl3 in H2SO4). "
            "IAA reacts to produce a pink/red colour measured at 530 nm. "
            "Used to quantify IAA production by PGPB isolates."
        ),
        "sample_data": {
            "concentration": [0, 5, 10, 20, 40, 60, 80, 100],
            "absorbance":    [0.000, 0.085, 0.172, 0.341, 0.669,
                              0.985, 1.274, 1.548]
        }
    },
    "Total Ammoniacal Nitrogen (TAN)": {
        "x_unit": "mg/L",
        "y_unit": "OD650",
        "wavelength": "650 nm",
        "model": "linear",
        "description": (
            "NH4-N standard curve at 650 nm. Used to quantify ammonium "
            "produced by nitrogen-fixing bacteria in culture media."
        ),
        "sample_data": {
            "concentration": [0, 0.5, 1.0, 2.0, 4.0, 6.0, 8.0, 10.0],
            "absorbance":    [0.000, 0.048, 0.097, 0.193, 0.381,
                              0.563, 0.739, 0.901]
        }
    },
    "Custom Assay": {
        "x_unit": "ug/mL",
        "y_unit": "OD",
        "wavelength": "custom",
        "model": "linear",
        "description": "Enter your own standard curve data.",
        "sample_data": {
            "concentration": [0, 10, 20, 40, 80, 100],
            "absorbance":    [0.000, 0.100, 0.200, 0.395, 0.780, 0.960]
        }
    }
}