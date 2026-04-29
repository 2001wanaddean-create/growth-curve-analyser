import numpy as np
from scipy.optimize import curve_fit
import plotly.graph_objects as go

def gompertz_model(t, A, mu_max, lam):
    return A * np.exp(-np.exp((mu_max * np.e / A) * (lam - t) + 1))

def fit_growth_curve(time, od600):
    try:
        A0   = max(od600) * 1.05
        mu0  = 0.3
        lam0 = time[np.argmax(np.gradient(od600, time))] * 0.5
        popt, pcov = curve_fit(
            gompertz_model, time, od600,
            p0=[A0, mu0, lam0],
            bounds=([0,0,0],[10,5,50]),
            maxfev=10000
        )
        A, mu_max, lam = popt
        return {"A":A,"mu_max":mu_max,"lam":lam,"success":True}
    except Exception as e:
        return {"success":False,"error":str(e)}

def detect_phases(time, od600, params):
    A, mu_max, lam = params["A"], params["mu_max"], params["lam"]
    fitted = gompertz_model(np.array(time), A, mu_max, lam)
    rates  = np.gradient(fitted, time)
    max_r  = np.argmax(rates)
    mask   = (rates < 0.1*rates[max_r]) & (time > time[max_r])
    log_end = time[mask][0] if mask.any() else time[-1]
    return {
        "lag_end":  round(float(lam),2),
        "log_start":round(float(lam),2),
        "log_end":  round(float(log_end),2),
        "stat_start":round(float(log_end),2)
    }

def calc_doubling_time(mu_max):
    return round(np.log(2)/mu_max, 2) if mu_max > 0 else None

def build_plotly_figure(time, od600, params, phases, organism):
    t_sm = np.linspace(min(time), max(time), 300)
    od_f = gompertz_model(t_sm, params["A"], params["mu_max"], params["lam"])
    fig  = go.Figure()
    regions = [
        ("Lag",        0,                    phases["lag_end"],   "rgba(255,200,80,0.13)"),
        ("Log",        phases["log_start"],  phases["log_end"],   "rgba(60,200,120,0.13)"),
        ("Stationary", phases["stat_start"], max(time),           "rgba(80,140,255,0.13)")
    ]
    for name, x0, x1, col in regions:
        fig.add_vrect(x0=x0, x1=x1, fillcolor=col, layer="below",
            line_width=0, annotation_text=f"{name}",
            annotation_position="top left", annotation_font_size=11)
    fig.add_trace(go.Scatter(x=time, y=od600, mode="markers",
        name="Observed OD₆₀₀",
        marker=dict(size=9,color="#1D9E75",line=dict(width=1.5,color="white"))))
    fig.add_trace(go.Scatter(x=t_sm, y=od_f, mode="lines",
        name="Gompertz fit", line=dict(color="#378ADD",width=2.5)))
    dt = calc_doubling_time(params["mu_max"])
    fig.update_layout(
        title=dict(text=f"Growth Curve — {organism}"
            f"μmax={params['mu_max']:.3f} h⁻¹  ·  "
            f"Doubling time={dt} h  ·  Lag={phases['lag_end']} h",
            font_size=15),
        xaxis_title="Time (hours)", yaxis_title="OD₆₀₀",
        plot_bgcolor="white", height=420,
        legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
        margin=dict(l=50,r=20,t=90,b=50)
    )
    fig.update_xaxes(showgrid=True,gridcolor="#f0f0f0")
    fig.update_yaxes(showgrid=True,gridcolor="#f0f0f0")
    return fig


def build_excel_growth_chart(time, od600, params, organism):
    import io
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, Reference
    from openpyxl.chart.series import SeriesFactory

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Observed Data"
    ws1["A1"], ws1["B1"] = "Time (hours)", "OD600 Observed"
    for i, (t, od) in enumerate(zip(time, od600), start=2):
        ws1.cell(i, 1, float(t))
        ws1.cell(i, 2, float(od))

    ws2 = wb.create_sheet("Gompertz Fit")
    t_fit = np.linspace(min(time), max(time), 100)
    od_fit = gompertz_model(t_fit, params["A"], params["mu_max"], params["lam"])
    ws2["A1"], ws2["B1"] = "Time (hours)", "OD600 Fitted"
    for i, (t, od) in enumerate(zip(t_fit, od_fit), start=2):
        ws2.cell(i, 1, float(t))
        ws2.cell(i, 2, float(od))

    ws3 = wb.create_sheet("Parameters")
    dt = round(float(np.log(2) / params["mu_max"]), 2) if params["mu_max"] > 0 else "N/A"
    for row in [
        ["Parameter", "Value"],
        ["Organism", organism],
        ["mu_max (h⁻¹)", round(float(params["mu_max"]), 4)],
        ["Asymptote A (OD)", round(float(params["A"]), 4)],
        ["Lag λ (hours)", round(float(params["lam"]), 4)],
        ["Doubling time (h)", dt],
    ]:
        ws3.append(row)

    n_obs = len(time)
    chart = ScatterChart()
    chart.title = f"Growth Curve — {organism}"
    chart.style = 10
    chart.x_axis.title = "Time (hours)"
    chart.y_axis.title = "OD600"
    chart.width, chart.height = 24, 14

    x_obs = Reference(ws1, min_col=1, min_row=2, max_row=n_obs + 1)
    y_obs = Reference(ws1, min_col=2, min_row=1, max_row=n_obs + 1)
    s_obs = SeriesFactory(y_obs, x_obs)
    s_obs.marker.symbol = "circle"
    s_obs.marker.size = 7
    s_obs.graphicalProperties.line.noFill = True
    chart.series.append(s_obs)

    x_fit = Reference(ws2, min_col=1, min_row=2, max_row=101)
    y_fit = Reference(ws2, min_col=2, min_row=1, max_row=101)
    s_fit = SeriesFactory(y_fit, x_fit)
    s_fit.marker.symbol = "none"
    s_fit.graphicalProperties.line.solidFill = "378ADD"
    chart.series.append(s_fit)

    ws1.add_chart(chart, "D2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()